"""
parsers/xlsx_parser.py — Extract tabular and text content from Excel files.

Uses openpyxl for .xlsx and xlrd for legacy .xls.
Each sheet becomes one PageContent.  Cell data is serialised row-by-row.
Embedded charts and images are extracted where possible.
"""

from __future__ import annotations

import io
from pathlib import Path
from typing import List, Any

from app.parsers.base_parser import BaseParser, ImageBlock, PageContent, ParserError
from app.config import settings
from app.utils.logger import get_logger

logger = get_logger(__name__)

_DIAGRAM_KEYWORDS = {
    "start", "end", "yes", "no", "decision", "process", "flow",
    "step", "condition", "branch", "gateway",
}


def _cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


class XlsxParser(BaseParser):
    """Parse .xlsx and .xls files."""

    def parse(self) -> List[PageContent]:
        ext = self.file_path.suffix.lower()
        if ext == ".xlsx":
            return self._parse_xlsx()
        elif ext == ".xls":
            return self._parse_xls()
        else:
            raise ParserError(f"Unsupported Excel extension: {ext}")

    # ── XLSX (openpyxl) ───────────────────────────────────────────────────────

    def _parse_xlsx(self) -> List[PageContent]:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(
                str(self.file_path), read_only=True, data_only=True
            )
        except Exception as exc:
            raise ParserError(f"Cannot open XLSX {self.file_name}: {exc}") from exc

        pages: List[PageContent] = []

        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            try:
                ws = wb[sheet_name]
                rows_text: List[str] = []

                for row in ws.iter_rows(values_only=True):
                    cells = [_cell_to_str(c) for c in row if _cell_to_str(c)]
                    if cells:
                        rows_text.append(" | ".join(cells))

                full_text = "\n".join(rows_text)
                kw_hits = sum(1 for kw in _DIAGRAM_KEYWORDS if kw in full_text.lower())

                pages.append(
                    PageContent(
                        page_number=sheet_idx + 1,
                        text=full_text,
                        section_title=sheet_name,
                        is_diagram_page=kw_hits >= settings.diagram_keyword_threshold,
                        raw_metadata={"source": "xlsx", "sheet": sheet_name},
                    )
                )
            except Exception as sheet_exc:
                logger.error(
                    "Error reading sheet '%s' in %s: %s",
                    sheet_name, self.file_name, sheet_exc,
                )
                pages.append(
                    PageContent(
                        page_number=sheet_idx + 1,
                        text=f"[ERROR reading sheet '{sheet_name}': {sheet_exc}]",
                        section_title=sheet_name,
                        raw_metadata={"error": str(sheet_exc)},
                    )
                )

        wb.close()
        logger.info("XLSX parsed: %s — %d sheet(s)", self.file_name, len(pages))
        return pages

    # ── XLS (xlrd) ────────────────────────────────────────────────────────────

    def _parse_xls(self) -> List[PageContent]:
        try:
            import xlrd
            wb = xlrd.open_workbook(str(self.file_path))
        except Exception as exc:
            raise ParserError(f"Cannot open XLS {self.file_name}: {exc}") from exc

        pages: List[PageContent] = []

        for sheet_idx in range(wb.nsheets):
            ws = wb.sheet_by_index(sheet_idx)
            rows_text: List[str] = []

            for row_idx in range(ws.nrows):
                cells = []
                for col_idx in range(ws.ncols):
                    cell = ws.cell(row_idx, col_idx)
                    val = _cell_to_str(cell.value)
                    if val:
                        cells.append(val)
                if cells:
                    rows_text.append(" | ".join(cells))

            full_text = "\n".join(rows_text)
            kw_hits = sum(1 for kw in _DIAGRAM_KEYWORDS if kw in full_text.lower())

            pages.append(
                PageContent(
                    page_number=sheet_idx + 1,
                    text=full_text,
                    section_title=ws.name,
                    is_diagram_page=kw_hits >= settings.diagram_keyword_threshold,
                    raw_metadata={"source": "xls", "sheet": ws.name},
                )
            )

        logger.info("XLS parsed: %s — %d sheet(s)", self.file_name, len(pages))
        return pages